#!/usr/bin/env python3
"""
Extract the "answer" field from each Hashtag API JSON response saved by
run_tests.sh (tests/<project>/q<N>.json) and compile everything into a
single spreadsheet: tests/answers.xlsx

It reads the PROJECTS and QUESTIONS arrays straight out of run_tests.sh so
the question text always lines up with q1.json, q2.json, etc. — no need to
retype the questions here.

Usage:
    python3 extract_answers.py [path/to/run_tests.sh] [path/to/results_dir] [output.xlsx]

All arguments are optional and default to:
    run_tests.sh   ./run_tests.sh
    results_dir    ./results
    output.xlsx    ./results/answers.xlsx
"""

import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def parse_bash_array(script_text: str, var_name: str) -> list[str]:
    """
    Pull a bash array like:
        QUESTIONS=(
          "question one"
          "question two (with parens)"
        )
    out of run_tests.sh.

    Parsed line-by-line rather than with a single greedy regex, because a
    question containing its own ")" (e.g. "...interlock loop (HVIL)...")
    would otherwise look like the end of the array and silently truncate
    everything after it.
    """
    lines = script_text.splitlines()
    items: list[str] = []
    in_array = False

    for raw_line in lines:
        stripped = raw_line.strip()

        if not in_array:
            if stripped.startswith(f"{var_name}=("):
                in_array = True
                # Handle a single-line array: VAR=("a" "b")
                if stripped.endswith(")") and stripped != f"{var_name}=(":
                    body = stripped[len(f"{var_name}=("):-1]
                    items.extend(re.findall(r'"([^"]*)"', body))
                    in_array = False
            continue

        # We're inside the array now
        if stripped == ")":
            in_array = False
            continue
        if not stripped or stripped.startswith("#"):
            continue
        items.extend(re.findall(r'"([^"]*)"', stripped))

    if not items:
        raise ValueError(f"Could not find array '{var_name}' in run_tests.sh")
    return items


def main():
    script_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("scripts/run_tests.sh")
    results_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("tests")
    output_path = Path(sys.argv[3]) if len(sys.argv) > 3 else results_dir / "answers.xlsx"

    if not script_path.exists():
        sys.exit(f"ERROR: {script_path} not found")
    if not results_dir.exists():
        sys.exit(f"ERROR: {results_dir} not found")

    script_text = script_path.read_text()
    projects = parse_bash_array(script_text, "PROJECTS")
    questions = parse_bash_array(script_text, "QUESTIONS")

    if not projects or not questions:
        sys.exit("ERROR: could not parse PROJECTS/QUESTIONS from run_tests.sh")

    rows = []  # (project, question, answer)
    for project in projects:
        for i, question in enumerate(questions, start=1):
            json_file = results_dir / project / f"q{i}.json"
            if not json_file.exists():
                answer = "[missing result file]"
            else:
                try:
                    data = json.loads(json_file.read_text())
                    answer = data.get("answer", "[no 'answer' field in response]")
                except json.JSONDecodeError:
                    answer = "[invalid JSON in response file]"
            rows.append((project, question, answer))

    # Build the spreadsheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Answers"

    headers = ["Project", "Question", "Answer"]
    ws.append(headers)
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(name="Arial", bold=True)

    for project, question, answer in rows:
        ws.append([project, question, answer])

    # Formatting: Arial font, wrap text, sensible column widths
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            if cell.row > 1:
                cell.font = Font(name="Arial")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Wrote {len(rows)} rows to {output_path}")


if __name__ == "__main__":
    main()